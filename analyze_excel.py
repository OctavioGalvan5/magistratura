import pandas as pd
from openpyxl import load_workbook

path = r"c:\Users\octav\Downloads\Consejo de la magistratura\AVALES_ELECCION_2026_control.xlsx"

wb = load_workbook(path, data_only=True)
print("=== SHEETS ===")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  - {name}: {ws.max_row} rows x {ws.max_column} cols")

print()
for name in wb.sheetnames:
    print(f"\n=== SHEET: {name} ===")
    df = pd.read_excel(path, sheet_name=name, header=None)
    # Show first 15 rows so we can spot headers
    print(df.head(15).to_string())
    print(f"\n  shape: {df.shape}")
